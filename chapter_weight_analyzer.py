#!/usr/bin/env python3
"""
Chapter Weight Analyzer - Smart Chapter Weight & Reading Planner
Analyzes bookmark distribution and generates reading schedules
"""

import os
import json
import csv
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta
import fitz  # PyMuPDF

# Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl.comments
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")


@dataclass
class Bookmark:
    """Bookmark data structure"""
    title: str
    page: int
    level: int = 1


@dataclass
class ChapterWeight:
    """Chapter weight analysis data"""
    title: str
    level: int
    start_page: int
    end_page: int
    page_count: int
    weight_percentage: float
    parent_title: Optional[str] = None


@dataclass
class ReadingPlanEntry:
    """Reading plan entry for a chapter"""
    chapter_title: str
    start_page: int
    end_page: int
    page_count: int
    weight_percentage: float
    assigned_days: int
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    # Optional enhancement: for merged chapters, store original titles for tooltip
    merged_chapter_titles: Optional[List[str]] = None


class ChapterWeightAnalyzer:
    """Analyzes chapter weights from bookmarks"""

    def __init__(self, bookmarks: List[Bookmark], total_pages: int):
        """
        Initialize analyzer with bookmarks and total pages

        Args:
            bookmarks: List of Bookmark objects
            total_pages: Total number of pages in the PDF
        """
        self.bookmarks = sorted(bookmarks, key=lambda b: b.page)
        self.total_pages = total_pages
        self.chapter_weights: List[ChapterWeight] = []

    def calculate_weights(self, include_level: int = 1) -> List[ChapterWeight]:
        """
        Calculate chapter weights based on page distribution

        Args:
            include_level: Maximum bookmark level to include (1 for main chapters only, 2+ for nested)

        Returns:
            List of ChapterWeight objects
        """
        self.chapter_weights = []

        # Filter bookmarks by level
        filtered_bookmarks = [b for b in self.bookmarks if b.level <= include_level]

        if not filtered_bookmarks:
            return []

        # Calculate weights for each bookmark
        for i, bookmark in enumerate(filtered_bookmarks):
            start_page = bookmark.page

            # Determine end page (next bookmark's start page - 1, or total pages)
            if i + 1 < len(filtered_bookmarks):
                end_page = filtered_bookmarks[i + 1].page - 1
            else:
                end_page = self.total_pages

            # Ensure end_page is not less than start_page - 1
            if end_page < start_page - 1:
                end_page = start_page - 1

            # Calculate page count and weight
            page_count = max(0, end_page - start_page + 1) # Ensure page_count is not negative
            if self.total_pages == 0:
                weight_percentage = 0.0
            else:
                weight_percentage = (page_count / self.total_pages) * 100

            # Find parent title for nested bookmarks
            parent_title = None
            if bookmark.level > 1:
                # Find the most recent level 1 bookmark
                for j in range(i - 1, -1, -1):
                    if filtered_bookmarks[j].level < bookmark.level:
                        parent_title = filtered_bookmarks[j].title
                        break

            chapter_weight = ChapterWeight(
                title=bookmark.title,
                level=bookmark.level,
                start_page=start_page,
                end_page=end_page,
                page_count=page_count,
                weight_percentage=weight_percentage,
                parent_title=parent_title
            )

            self.chapter_weights.append(chapter_weight)

        return self.chapter_weights

    def get_level1_weights(self) -> List[ChapterWeight]:
        """Get weights for level 1 (main chapters) only"""
        return self.calculate_weights(include_level=1)

    def get_hierarchical_weights(self, max_level: int = 2) -> List[ChapterWeight]:
        """Get weights including nested chapters"""
        return self.calculate_weights(include_level=max_level)

    def get_statistics(self) -> Dict[str, Any]:
        """Get statistical summary of chapter weights"""
        if not self.chapter_weights:
            return {}

        page_counts = [cw.page_count for cw in self.chapter_weights if cw.page_count > 0] # Filter out 0-page chapters
        weights = [cw.weight_percentage for cw in self.chapter_weights if cw.page_count > 0] # Filter out 0-page chapters

        if not page_counts: # If no valid chapters after filtering
            return {
                'total_chapters': 0,
                'total_pages': self.total_pages,
                'average_pages_per_chapter': 0.0,
                'average_weight_percentage': 0.0,
                'longest_chapter': {'title': 'N/A', 'pages': 0, 'percentage': 0.0},
                'shortest_chapter': {'title': 'N/A', 'pages': 0, 'percentage': 0.0},
                'top_5_largest': []
            }

        # Find longest and shortest chapters (from non-zero page chapters)
        longest_chapter = max([cw for cw in self.chapter_weights if cw.page_count > 0], key=lambda cw: cw.page_count)
        shortest_chapter = min([cw for cw in self.chapter_weights if cw.page_count > 0], key=lambda cw: cw.page_count)

        # Get top 5 largest chapters
        top_5_largest = sorted([cw for cw in self.chapter_weights if cw.page_count > 0], key=lambda cw: cw.page_count, reverse=True)[:5]

        return {
            'total_chapters': len(self.chapter_weights),
            'total_pages': self.total_pages,
            'average_pages_per_chapter': sum(page_counts) / len(page_counts),
            'average_weight_percentage': sum(weights) / len(weights),
            'longest_chapter': {
                'title': longest_chapter.title,
                'pages': longest_chapter.page_count,
                'percentage': longest_chapter.weight_percentage
            },
            'shortest_chapter': {
                'title': shortest_chapter.title,
                'pages': shortest_chapter.page_count,
                'percentage': shortest_chapter.weight_percentage
            },
            'top_5_largest': [
                {
                    'title': cw.title,
                    'pages': cw.page_count,
                    'percentage': cw.weight_percentage
                }
                for cw in top_5_largest
            ]
        }

    def export_to_dict(self) -> List[Dict[str, Any]]:
        """Export chapter weights as list of dictionaries"""
        return [asdict(cw) for cw in self.chapter_weights]

    def export_to_json(self, filepath: str) -> bool:
        """Export chapter weights to JSON file"""
        try:
            data = {
                'total_pages': self.total_pages,
                'analysis_date': datetime.now().isoformat(),
                'chapters': self.export_to_dict(),
                'statistics': self.get_statistics()
            }

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True
        except Exception as e:
            print(f"Error exporting to JSON: {e}")
            return False

    def export_to_csv(self, filepath: str) -> bool:
        """Export chapter weights to CSV file"""
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                if not self.chapter_weights:
                    return False

                fieldnames = ['title', 'level', 'start_page', 'end_page',
                             'page_count', 'weight_percentage', 'parent_title']
                writer = csv.DictWriter(f, fieldnames=fieldnames)

                writer.writeheader()
                for cw in self.chapter_weights:
                    writer.writerow(asdict(cw))

            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False

    def export_to_excel(self, filepath: str) -> bool:
        """
        Export chapter weights to professional Excel file

        Args:
            filepath: Path to save Excel file

        Returns:
            True if successful, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl is required for Excel export")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Chapter Weights"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            cell_alignment = Alignment(horizontal="left", vertical="center")
            number_alignment = Alignment(horizontal="right", vertical="center")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ["Chapter Title", "Level", "Start Page", "End Page", "Page Count", "Weight %"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Data rows
            for row_num, cw in enumerate(self.chapter_weights, 2):
                ws.cell(row=row_num, column=1, value=cw.title).alignment = cell_alignment
                ws.cell(row=row_num, column=2, value=cw.level).alignment = number_alignment
                ws.cell(row=row_num, column=3, value=cw.start_page).alignment = number_alignment
                ws.cell(row=row_num, column=4, value=cw.end_page).alignment = number_alignment
                ws.cell(row=row_num, column=5, value=cw.page_count).alignment = number_alignment
                ws.cell(row=row_num, column=6, value=f"{cw.weight_percentage:.2f}%").alignment = number_alignment

                # Apply borders
                for col_num in range(1, 7):
                    ws.cell(row=row_num, column=col_num).border = border

            # Add statistics sheet
            stats_ws = wb.create_sheet("Statistics")
            stats_ws.cell(row=1, column=1, value="Statistic").font = header_font
            stats_ws.cell(row=1, column=1).fill = header_fill
            stats_ws.cell(row=1, column=2, value="Value").font = header_font
            stats_ws.cell(row=1, column=2).fill = header_fill

            stats = self.get_statistics()
            row = 2
            for key, value in stats.items():
                # Handle nested dictionaries for longest/shortest chapter
                if isinstance(value, dict):
                    stats_ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                    stats_ws.cell(row=row, column=2, value=f"{value.get('title', 'N/A')} ({value.get('pages', 0)} pages, {value.get('percentage', 0):.1f}%)")
                elif isinstance(value, list):
                    stats_ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                    stats_ws.cell(row=row, column=2, value="See below for details")
                    # Add detailed list for top_5_largest
                    for item in value:
                        row += 1
                        stats_ws.cell(row=row, column=1, value=f"- {item.get('title', 'N/A')}")
                        stats_ws.cell(row=row, column=2, value=f"{item.get('pages', 0)} pages ({item.get('percentage', 0):.1f}%)")
                else:
                    stats_ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
                    stats_ws.cell(row=row, column=2, value=str(value))
                row += 1

            # Adjust column widths
            for ws_sheet in [ws, stats_ws]:
                for column in ws_sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 75) # Increased max width for longer titles/comments
                    ws_sheet.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    def export_to_markdown(self, filepath: str) -> bool:
        """
        Export chapter weights to Obsidian-compatible Markdown

        Args:
            filepath: Path to save Markdown file

        Returns:
            True if successful, False otherwise
        """
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                # Front matter for Obsidian
                f.write("---\n")
                f.write("tags: [reading, analysis, chapters]\n")
                f.write(f"created: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write("type: chapter-analysis\n")
                f.write("---\n\n")

                # Title
                f.write("# 📊 Chapter Weight Analysis\n\n")

                # Statistics
                stats = self.get_statistics()
                f.write("## 📈 Statistics\n\n")
                f.write(f"- **Total Chapters**: {stats.get('total_chapters', 0)}\n")
                f.write(f"- **Total Pages**: {stats.get('total_pages', 0)}\n")
                f.write(f"- **Average Chapter Length**: {stats.get('average_pages_per_chapter', 0):.1f} pages\n")
                f.write(f"- **Longest Chapter**: {stats.get('longest_chapter', {}).get('title', 'N/A')} ({stats.get('longest_chapter', {}).get('pages', 0)} pages)\n")
                f.write(f"- **Shortest Chapter**: {stats.get('shortest_chapter', {}).get('title', 'N/A')} ({stats.get('shortest_chapter', {}).get('pages', 0)} pages)\n\n")

                # Chapter weights table
                f.write("## 📚 Chapter Weights\n\n")
                f.write("| Chapter Title | Level | Pages | Start | End | Weight % |\n")
                f.write("|---------------|-------|-------|-------|-----|----------|\n")

                for cw in self.chapter_weights:
                    # Escape pipe characters in titles
                    title = cw.title.replace('|', '\\|')
                    f.write(f"| {title} | {cw.level} | {cw.page_count} | {cw.start_page} | {cw.end_page} | {cw.weight_percentage:.2f}% |\n")

                f.write("\n")

                # Top 5 longest chapters
                f.write("## 🏆 Top 5 Longest Chapters\n\n")
                sorted_chapters = sorted([cw for cw in self.chapter_weights if cw.page_count > 0], key=lambda x: x.page_count, reverse=True)[:5]
                for i, cw in enumerate(sorted_chapters, 1):
                    f.write(f"{i}. **{cw.title}** - {cw.page_count} pages ({cw.weight_percentage:.1f}%)\n")

                f.write("\n---\n")
                f.write(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")

            return True
        except PermissionError as e:
            print(f"Permission denied when exporting to Markdown: {e}")
            raise PermissionError(f"Cannot write to file. Please check:\n1. File is not open in another program\n2. You have write permissions\n3. The directory exists")
        except OSError as e:
            print(f"OS error when exporting to Markdown: {e}")
            raise OSError(f"File system error: {str(e)}")
        except Exception as e:
            print(f"Error exporting to Markdown: {e}")
            raise Exception(f"Failed to export markdown: {type(e).__name__} - {str(e)}")


class ReadingPlanGenerator:
    """Generates reading schedules based on chapter weights"""

    def __init__(self, chapter_weights: List[ChapterWeight], total_target_days: int,
                 start_date: Optional[datetime] = None, weekends_off: bool = False, weekend_days: Optional[List[int]] = None):
        """
        Initialize reading plan generator

        Args:
            chapter_weights: List of ChapterWeight objects
            total_target_days: The desired total number of days for the reading plan.
            start_date: Optional start date (defaults to today)
            weekends_off: If True, weekend days will not be assigned reading.
            weekend_days: List of weekday indices to skip (0=Monday, 6=Sunday). Defaults to [5, 6] (Sat, Sun).
        """
        # Filter out chapters with 0 pages before processing
        self.chapter_weights = [cw for cw in chapter_weights if cw.page_count > 0]
        self.total_target_days = total_target_days
        self.start_date = start_date or datetime.now()
        self.weekends_off = weekends_off
        self.weekend_days = weekend_days if weekend_days is not None else [5, 6]  # Default to Saturday and Sunday
        self.reading_plan: List[ReadingPlanEntry] = []

    def _calculate_proportional_days(self) -> List[Dict[str, Any]]:
        """
        Calculate proportional days for each chapter based on weight percentage.

        Returns:
            List of dicts with 'chapter', 'raw_days', and 'rounded_days'
        """
        return [
            {
                'chapter': cw,
                'raw_days': (cw.weight_percentage / 100.0) * self.total_target_days,
                'rounded_days': round((cw.weight_percentage / 100.0) * self.total_target_days)
            }
            for cw in self.chapter_weights
        ]

    def _merge_zero_day_chapters(self, chapter_day_allocations: List[Dict[str, Any]]) -> List[Tuple[List[ChapterWeight], int]]:
        """
        Merge chapters with 0 allocated days into adjacent chapters.

        Args:
            chapter_day_allocations: List of chapter allocations with days

        Returns:
            List of tuples (chapter_group, days) representing merged chapters
        """
        processed_items: List[Tuple[List[ChapterWeight], int]] = []
        i = 0

        while i < len(chapter_day_allocations):
            current_alloc = chapter_day_allocations[i]

            if current_alloc['rounded_days'] == 0:
                merge_direction = self._determine_merge_direction(chapter_day_allocations, i)

                if merge_direction == 'backward' and processed_items:
                    # Merge into previous group
                    last_group_chapters, last_group_days = processed_items[-1]
                    last_group_chapters.append(current_alloc['chapter'])
                    processed_items[-1] = (last_group_chapters, last_group_days)
                    i += 1
                else:  # merge forward
                    group, next_idx = self._merge_forward(chapter_day_allocations, i)
                    processed_items.append(group)
                    i = next_idx
            else:
                # Chapter has days > 0, keep it standalone
                processed_items.append(([current_alloc['chapter']], current_alloc['rounded_days']))
                i += 1

        return processed_items

    def _determine_merge_direction(self, allocations: List[Dict[str, Any]], index: int) -> str:
        """Determine whether to merge a zero-day chapter backward or forward."""
        if index == 0:
            return 'forward'
        if index == len(allocations) - 1:
            return 'backward'

        prev_weight = allocations[index - 1]['chapter'].weight_percentage
        next_weight = allocations[index + 1]['chapter'].weight_percentage

        return 'backward' if prev_weight <= next_weight else 'forward'

    def _merge_forward(self, allocations: List[Dict[str, Any]], start_idx: int) -> Tuple[Tuple[List[ChapterWeight], int], int]:
        """
        Merge current chapter forward with subsequent chapters.

        Returns:
            Tuple of ((chapter_group, days), next_index)
        """
        group = [allocations[start_idx]['chapter']]
        j = start_idx + 1

        # Add all consecutive zero-day chapters
        while j < len(allocations) and allocations[j]['rounded_days'] == 0:
            group.append(allocations[j]['chapter'])
            j += 1

        # Add the first non-zero chapter or assign 1 day if all remaining are zero
        if j < len(allocations):
            group.append(allocations[j]['chapter'])
            days = max(1, allocations[j]['rounded_days'])
            return ((group, days), j + 1)
        else:
            return ((group, 1), j)

    def _adjust_days_to_target(self, processed_items: List[Tuple[List[ChapterWeight], int]],
                               assigned_days_list: List[int]) -> List[int]:
        """
        Adjust assigned days to match total_target_days exactly.

        Args:
            processed_items: List of (chapter_group, days) tuples
            assigned_days_list: Current day assignments

        Returns:
            Adjusted list of assigned days
        """
        max_iterations = len(assigned_days_list) * 2
        iter_count = 0

        while sum(assigned_days_list) != self.total_target_days and iter_count < max_iterations:
            difference = self.total_target_days - sum(assigned_days_list)

            if difference > 0:
                # Add days to highest weight groups
                idx_to_add = self._find_highest_weight_group(processed_items)
                if idx_to_add is not None:
                    assigned_days_list[idx_to_add] += 1
            else:
                # Remove days from lowest weight groups (keeping minimum 1 day)
                idx_to_remove = self._find_lowest_weight_group_with_extra_days(processed_items, assigned_days_list)
                if idx_to_remove is not None:
                    assigned_days_list[idx_to_remove] -= 1
                else:
                    break  # All groups have 1 day, cannot remove further

            iter_count += 1

        # Final adjustment for any remaining difference
        return self._final_day_adjustment(assigned_days_list)

    def _find_highest_weight_group(self, processed_items: List[Tuple[List[ChapterWeight], int]]) -> Optional[int]:
        """Find the index of the group with the highest total weight."""
        if not processed_items:
            return None

        group_weights = [
            (idx, sum(ch.weight_percentage for ch in group_chapters))
            for idx, (group_chapters, _) in enumerate(processed_items)
        ]
        return max(group_weights, key=lambda x: x[1])[0] if group_weights else None

    def _find_lowest_weight_group_with_extra_days(self, processed_items: List[Tuple[List[ChapterWeight], int]],
                                                   assigned_days_list: List[int]) -> Optional[int]:
        """Find the index of the lowest weight group that has more than 1 day."""
        group_weights = [
            (idx, sum(ch.weight_percentage for ch in group_chapters))
            for idx, (group_chapters, _) in enumerate(processed_items)
            if assigned_days_list[idx] > 1
        ]
        return min(group_weights, key=lambda x: x[1])[0] if group_weights else None

    def _final_day_adjustment(self, assigned_days_list: List[int]) -> List[int]:
        """Make final adjustments to match target days exactly."""
        final_diff = self.total_target_days - sum(assigned_days_list)

        if final_diff > 0:
            for i in range(final_diff):
                if assigned_days_list:
                    assigned_days_list[i % len(assigned_days_list)] += 1
        elif final_diff < 0:
            for i in range(abs(final_diff)):
                if assigned_days_list:
                    idx = i % len(assigned_days_list)
                    if assigned_days_list[idx] > 1:
                        assigned_days_list[idx] -= 1

        return assigned_days_list

    def _create_chapter_group_title(self, chapters_group: List[ChapterWeight]) -> str:
        """Create a display title for a group of chapters."""
        if len(chapters_group) == 1:
            return chapters_group[0].title
        elif len(chapters_group) == 2:
            return f"{chapters_group[0].title} – {chapters_group[1].title}"
        else:
            return f"{chapters_group[0].title} – {chapters_group[-1].title}"

    def _calculate_end_date(self, start_date: datetime, days: int) -> datetime:
        """Calculate end date respecting weekend settings."""
        actual_reading_days_counter = 0
        temp_end_date = start_date

        while actual_reading_days_counter < days:
            if not self.weekends_off or temp_end_date.weekday() not in self.weekend_days:
                actual_reading_days_counter += 1
            if actual_reading_days_counter < days:
                temp_end_date += timedelta(days=1)

        return temp_end_date

    def _skip_to_next_reading_day(self, date: datetime) -> datetime:
        """Skip to the next valid reading day (non-weekend if weekends_off is True)."""
        while self.weekends_off and date.weekday() in self.weekend_days:
            date += timedelta(days=1)
        return date

    def generate_plan(self) -> List[ReadingPlanEntry]:
        """
        Generate intelligent reading plan with proportional day allocation based on chapter weights.

        Algorithm:
        1. Calculate raw days for each chapter: raw_days = weight_percentage × total_days
        2. Round raw days to nearest integer
        3. Handle zero-day chapters by merging with adjacent chapters based on weight comparison
        4. Ensure every chapter/group gets at least 1 day
        5. Adjust total to match target days exactly
        6. Assign actual start/end dates, skipping weekends if `weekends_off` is True

        Returns:
            List of ReadingPlanEntry objects
        """
        self.reading_plan = []

        if not self.chapter_weights or self.total_target_days <= 0:
            return []

        total_pages = sum(cw.page_count for cw in self.chapter_weights)
        if total_pages == 0:
            return []

        # Step 1: Calculate proportional days
        chapter_day_allocations = self._calculate_proportional_days()

        # Step 2: Merge zero-day chapters
        processed_items = self._merge_zero_day_chapters(chapter_day_allocations)

        # Step 3: Ensure minimum 1 day for all groups
        assigned_days_list = [max(1, days) for _, days in processed_items]

        # Step 4: Adjust to match total_target_days exactly
        assigned_days_list = self._adjust_days_to_target(processed_items, assigned_days_list)

        # Step 5: Generate plan entries with dates
        current_date = self.start_date

        for (chapters_group, _), days in zip(processed_items, assigned_days_list):
            if days == 0:
                continue

            # Skip to next reading day
            current_date = self._skip_to_next_reading_day(current_date)

            # Create entry
            title = self._create_chapter_group_title(chapters_group)
            start_date_str = current_date.strftime("%Y-%m-%d")
            end_date = self._calculate_end_date(current_date, days)
            end_date_str = end_date.strftime("%Y-%m-%d")

            entry = ReadingPlanEntry(
                chapter_title=title,
                start_page=chapters_group[0].start_page,
                end_page=chapters_group[-1].end_page,
                page_count=sum(ch.page_count for ch in chapters_group),
                weight_percentage=sum(ch.weight_percentage for ch in chapters_group),
                assigned_days=days,
                start_date=start_date_str,
                end_date=end_date_str,
                merged_chapter_titles=[ch.title for ch in chapters_group] if len(chapters_group) > 1 else None
            )

            self.reading_plan.append(entry)
            current_date = end_date + timedelta(days=1)

        return self.reading_plan

    def get_daily_pages(self) -> float:
        """Calculate average daily pages for the plan, considering only assigned reading days."""
        if not self.reading_plan:
            return 0.0

        total_pages = sum(entry.page_count for entry in self.reading_plan)
        actual_assigned_days = sum(entry.assigned_days for entry in self.reading_plan)

        if actual_assigned_days == 0:
            return 0.0

        return total_pages / actual_assigned_days

    def export_to_dict(self) -> List[Dict[str, Any]]:
        """Export reading plan as list of dictionaries"""
        return [asdict(entry) for entry in self.reading_plan]

    def export_to_json(self, filepath: str) -> bool:
        """Export reading plan to JSON file"""
        try:
            data = {
                'total_target_days': self.total_target_days,
                'actual_assigned_days': sum(entry.assigned_days for entry in self.reading_plan),
                'start_date': self.start_date.strftime("%Y-%m-%d"),
                'average_daily_pages': self.get_daily_pages(),
                'plan_created': datetime.now().isoformat(),
                'reading_plan': self.export_to_dict()
            }

            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True
        except Exception as e:
            print(f"Error exporting reading plan to JSON: {e}")
            return False

    def export_to_csv(self, filepath: str) -> bool:
        """Export reading plan to CSV file"""
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                if not self.reading_plan:
                    return False

                fieldnames = ['chapter_title', 'start_page', 'end_page', 'page_count',
                             'weight_percentage', 'assigned_days', 'start_date', 'end_date', 'merged_chapter_titles']
                writer = csv.DictWriter(f, fieldnames=fieldnames)

                writer.writeheader()
                for entry in self.reading_plan:
                    # Convert list of merged_chapter_titles to a string for CSV
                    entry_dict = asdict(entry)
                    if entry_dict['merged_chapter_titles'] is not None:
                        entry_dict['merged_chapter_titles'] = "; ".join(entry_dict['merged_chapter_titles'])
                    writer.writerow(entry_dict)

            return True
        except Exception as e:
            print(f"Error exporting reading plan to CSV: {e}")
            return False

    def export_to_text(self, filepath: str) -> bool:
        """Export reading plan to formatted text file"""
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("READING PLAN\n")
                f.write("=" * 80 + "\n\n")

                f.write(f"Total Target Duration: {self.total_target_days} days\n")
                f.write(f"Actual Assigned Duration: {sum(entry.assigned_days for entry in self.reading_plan)} days\n")
                f.write(f"Start Date: {self.start_date.strftime('%Y-%m-%d')}\n")
                f.write(f"Average Daily Pages (Active Days): {self.get_daily_pages():.1f}\n\n")

                f.write("-" * 80 + "\n\n")

                current_reading_day_counter = 1
                for i, entry in enumerate(self.reading_plan, 1):
                    f.write(f"Block {i} (Days {current_reading_day_counter}-{current_reading_day_counter + entry.assigned_days - 1}) - {entry.start_date} to {entry.end_date}\n")
                    f.write(f"Chapter: {entry.chapter_title}\n")
                    if entry.merged_chapter_titles:
                        f.write(f"  (Includes: {', '.join(entry.merged_chapter_titles)})\n")
                    f.write(f"Pages: {entry.start_page}-{entry.end_page} ({entry.page_count} pages)\n")
                    f.write(f"Duration: {entry.assigned_days} day(s)\n")
                    f.write(f"Weight: {entry.weight_percentage:.2f}%\n")
                    f.write("\n")
                    current_reading_day_counter += entry.assigned_days

                f.write("=" * 80 + "\n")
                f.write(f"Plan created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            return True
        except Exception as e:
            print(f"Error exporting reading plan to text: {e}")
            return False

    def export_to_excel(self, filepath: str) -> bool:
        """
        Export reading plan to professional Excel file

        Args:
            filepath: Path to save Excel file

        Returns:
            True if successful, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl is required for Excel export")
            return False

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reading Plan"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            title_font = Font(bold=True, size=14, color="1F4E78")

            cell_alignment = Alignment(horizontal="left", vertical="center")
            number_alignment = Alignment(horizontal="right", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = "📚 Reading Plan"
            title_cell.font = title_font
            title_cell.alignment = center_alignment

            # Summary info
            ws.merge_cells('A2:B2')
            ws['A2'] = "Total Target Duration:"
            ws['C2'] = f"{self.total_target_days} days"

            ws.merge_cells('A3:B3')
            ws['A3'] = "Actual Assigned Duration:"
            ws['C3'] = f"{sum(entry.assigned_days for entry in self.reading_plan)} days"

            ws.merge_cells('A4:B4')
            ws['A4'] = "Start Date:"
            ws['C4'] = self.start_date.strftime('%Y-%m-%d')

            ws.merge_cells('A5:B5')
            ws['A5'] = "Average Daily Pages (Active Days):"
            ws['C5'] = f"{self.get_daily_pages():.1f}"

            # Headers (row 7)
            headers = ["Chapter Title", "Start Page", "End Page", "Page Count", "Weight %", "Days", "Start Date", "End Date"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Data rows
            for row_num, entry in enumerate(self.reading_plan, 8):
                title_cell = ws.cell(row=row_num, column=1, value=entry.chapter_title)
                title_cell.alignment = cell_alignment
                if entry.merged_chapter_titles:
                    # Add tooltip for merged chapters
                    title_cell.comment = openpyxl.comments.Comment(
                        f"Includes: {', '.join(entry.merged_chapter_titles)}\n"
                        f"Total: {entry.page_count} pages, Duration: {entry.assigned_days} days",
                        author="ReadingPlanGenerator"
                    )

                ws.cell(row=row_num, column=2, value=entry.start_page).alignment = number_alignment
                ws.cell(row=row_num, column=3, value=entry.end_page).alignment = number_alignment
                ws.cell(row=row_num, column=4, value=entry.page_count).alignment = number_alignment
                ws.cell(row=row_num, column=5, value=f"{entry.weight_percentage:.2f}%").alignment = number_alignment
                ws.cell(row=row_num, column=6, value=entry.assigned_days).alignment = number_alignment
                ws.cell(row=row_num, column=7, value=entry.start_date).alignment = center_alignment
                ws.cell(row=row_num, column=8, value=entry.end_date).alignment = center_alignment

                # Apply borders
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).border = border

            # Adjust column widths
            column_widths = [40, 12, 12, 12, 10, 8, 12, 12]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Freeze header row
            ws.freeze_panes = 'A8'

            wb.save(filepath)
            return True

        except Exception as e:
            print(f"Error exporting reading plan to Excel: {e}")
            return False

    def export_to_markdown(self, filepath: str) -> bool:
        """
        Export reading plan to Obsidian-compatible Markdown

        Args:
            filepath: Path to save Markdown file

        Returns:
            True if successful, False otherwise
        """
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                # Front matter for Obsidian
                f.write("---\n")
                f.write("tags: [reading, plan, schedule]\n")
                f.write(f"created: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write(f"start_date: {self.start_date.strftime('%Y-%m-%d')}\n")
                f.write(f"duration_target: {self.total_target_days}\n")
                f.write(f"duration_actual: {sum(entry.assigned_days for entry in self.reading_plan)}\n")
                f.write("type: reading-plan\n")
                f.write("---\n\n")

                # Title
                f.write("# 📚 Reading Plan\n\n")

                # Summary
                f.write("## 📋 Plan Summary\n\n")
                f.write(f"- **Total Target Duration**: {self.total_target_days} days\n")
                f.write(f"- **Actual Assigned Duration**: {sum(entry.assigned_days for entry in self.reading_plan)} days\n")
                f.write(f"- **Start Date**: {self.start_date.strftime('%Y-%m-%d')}\n")

                if self.reading_plan:
                    last_entry = self.reading_plan[-1]
                    f.write(f"- **End Date**: {last_entry.end_date}\n")

                f.write(f"- **Average Daily Pages (Active Days)**: {self.get_daily_pages():.1f}\n")
                f.write(f"- **Total Reading Blocks**: {len(self.reading_plan)}\n\n")

                # Reading schedule
                f.write("## 📅 Reading Schedule\n\n")

                current_day_overall = 1
                for entry in self.reading_plan:
                    f.write(f"### {entry.chapter_title}\n\n")
                    if entry.merged_chapter_titles:
                        f.write(f"  > **Includes**: {', '.join(entry.merged_chapter_titles)}\n\n")
                    f.write(f"- **Pages**: {entry.start_page}-{entry.end_page} ({entry.page_count} pages)\n")
                    f.write(f"- **Duration**: {entry.assigned_days} day(s)\n")
                    f.write(f"- **Dates**: {entry.start_date} → {entry.end_date}\n")
                    f.write(f"- **Weight**: {entry.weight_percentage:.2f}%\n")

                    # Add checkboxes for each day within this block
                    f.write(f"\n**Daily Progress (Reading Days {current_day_overall} to {current_day_overall + entry.assigned_days - 1})**:\n")
                    for day_offset in range(entry.assigned_days):
                        # Calculate the actual calendar date for this specific reading day within the block
                        # This logic needs to correctly skip weekends for individual days in a block too.
                        temp_date_for_checkbox = datetime.strptime(entry.start_date, '%Y-%m-%d')
                        days_skipped_for_offset = 0
                        for d_off in range(day_offset):
                            temp_date_for_checkbox += timedelta(days=1)
                            while self.weekends_off and temp_date_for_checkbox.weekday() in self.weekend_days:
                                temp_date_for_checkbox += timedelta(days=1)

                        day_date_for_checkbox = temp_date_for_checkbox.strftime('%Y-%m-%d')
                        f.write(f"- [ ] Day {current_day_overall + day_offset} ({day_date_for_checkbox})\n")
                    current_day_overall += entry.assigned_days
                    f.write("\n")

                # Quick reference table
                f.write("## 📊 Quick Reference\n\n")
                f.write("| Reading Block | Pages | Days | Start Date | End Date |\n")
                f.write("|:--------------|-----:|----:|:----------:|:--------:|\n")
                lines = []
                for entry in self.reading_plan:
                    title = entry.chapter_title.replace('|', '\\|').replace('\n', ' ')
                    lines.append(f"| {title} | {entry.page_count} | {entry.assigned_days} | {entry.start_date} | {entry.end_date} |")
                f.write("\n".join(lines))
                f.write("\n\n---\n")
                f.write(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")

            return True
        except PermissionError as e:
            print(f"Permission denied when exporting reading plan to Markdown: {e}")
            raise PermissionError(f"Cannot write to file. Please check:\n1. File is not open in another program\n2. You have write permissions\n3. The directory exists")
        except OSError as e:
            print(f"OS error when exporting reading plan to Markdown: {e}")
            raise OSError(f"File system error: {str(e)}")
        except Exception as e:
            print(f"Error exporting reading plan to Markdown: {e}")
            raise Exception(f"Failed to export markdown: {type(e).__name__} - {str(e)}")


    def export_to_obsidian_markdown(self, filepath: str) -> bool:
        """Export reading plan to Obsidian-friendly Markdown (compact spacing, proper tables)."""
        try:
            schedule = self.reading_plan
            if not schedule:
                return False

            # Compute basic stats
            total_days = sum(s.assigned_days for s in schedule)
            total_pages = sum(s.page_count for s in schedule)
            avg_pages = round(total_pages / total_days, 2) if total_days else 0
            total_sessions = len(schedule)

            with open(filepath, 'w', encoding='utf-8') as f:
                # Front matter
                f.write("---\n")
                f.write("tags: [reading, plan, schedule]\n")
                f.write(f"created: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write("type: reading-plan\n")
                f.write("---\n\n")

                # Title
                f.write("# Reading Plan\n\n")

                # Summary
                f.write("## Plan Summary\n\n")
                f.write(f"- **Total Days**: {total_days}\n")
                f.write(f"- **Total Pages**: {total_pages}\n")
                f.write(f"- **Daily Average**: {avg_pages} pages\n")
                f.write(f"- **Total Blocks**: {total_sessions}\n\n")
                f.write("---\n\n")

                # Schedule
                f.write("## Reading Schedule\n\n")
                for i, session in enumerate(schedule, 1):
                    f.write(f"### Block {i}: {session.chapter_title}\n\n")
                    f.write(f"**Pages**: {session.start_page}-{session.end_page} ({session.page_count} pages)  \n")
                    f.write(f"**Duration**: {session.assigned_days} day{'s' if session.assigned_days > 1 else ''}  \n")
                    f.write(f"**Dates**: {session.start_date} → {session.end_date}  \n\n")

                    # Daily checkboxes
                    f.write("**Daily Progress**:\n")
                    current = datetime.strptime(session.start_date, "%Y-%m-%d")
                    end = datetime.strptime(session.end_date, "%Y-%m-%d")
                    day_num = sum(s.assigned_days for s in schedule[:i-1]) + 1
                    while current <= end:
                        f.write(f"- [ ] Day {day_num} ({current.strftime('%Y-%m-%d')})\n")
                        day_num += 1
                        current += timedelta(days=1)
                    f.write("\n---\n\n")

                # Quick table
                f.write("## Quick Reference\n\n")
                f.write("| Block | Pages | Days | Start | End |\n")
                f.write("|:------|------:|-----:|:-----:|:---:|\n")
                for i, s in enumerate(schedule, 1):
                    title = s.chapter_title[:30] + "..." if len(s.chapter_title) > 30 else s.chapter_title
                    start = datetime.strptime(s.start_date, "%Y-%m-%d").strftime("%b %d")
                    end = datetime.strptime(s.end_date, "%Y-%m-%d").strftime("%b %d")
                    f.write(f"| {title} | {s.page_count} | {s.assigned_days} | {start} | {end} |\n")

            return True
        except Exception as e:
            print(f"Error exporting to Obsidian markdown: {e}")
            return False


# --- Utility functions for PDF parsing (as provided in the original code) ---

def extract_bookmarks_from_pdf(pdf_path: str) -> List[Bookmark]:
    """
    Extracts bookmarks (table of contents) from a PDF file.

    Args:
        pdf_path: Path to the PDF file.

    Returns:
        A list of Bookmark objects.
    """
    bookmarks: List[Bookmark] = []
    try:
        doc = fitz.open(pdf_path)
        toc = doc.get_toc()
        for item in toc:
            level, title, page = item[0], item[1], item[2]
            # PyMuPDF pages are 1-indexed in TOC
            bookmarks.append(Bookmark(title=title, page=page))
        doc.close()
    except Exception as e:
        print(f"Error extracting bookmarks from PDF: {e}")
    return bookmarks

def get_total_pdf_pages(pdf_path: str) -> int:
    """
    Gets the total number of pages in a PDF file.

    Args:
        pdf_path: Path to the PDF file.

    Returns:
        Total number of pages, or 0 if an error occurs.
    """
    try:
        doc = fitz.open(pdf_path)
        total_pages = doc.page_count
        doc.close()
        return total_pages
    except Exception as e:
        print(f"Error getting total PDF pages: {e}")
        return 0

# --- Main execution example (simplified for demonstration) ---
if __name__ == "__main__":
    # --- Example Usage with placeholder data ---
    # In a real application, replace this with dynamic PDF parsing:
    # pdf_file = "your_book.pdf"
    # bookmarks = extract_bookmarks_from_pdf(pdf_file)
    # total_pages = get_total_pdf_pages(pdf_file)

    # Placeholder bookmarks, ensure the last page has a bookmark to delimit the end chapter
    # (or adjust ChapterWeightAnalyzer to handle the final pages if no last bookmark exists)
    placeholder_bookmarks = [
        Bookmark("Chapter 1: Intro", 1),
        Bookmark("Chapter 1.1: Sub-Intro A", 5),
        Bookmark("Chapter 1.2: Sub-Intro B", 7),
        Bookmark("Chapter 2: Core Concepts", 15),
        Bookmark("Chapter 2.1: Key Principles", 18),
        Bookmark("Chapter 3: Advanced Topics", 40),
        Bookmark("Chapter 3.1: Topic X", 42),
        Bookmark("Chapter 3.2: Topic Y", 48),
        Bookmark("Chapter 3.3: Topic Z", 50),
        Bookmark("Chapter 4: Case Studies", 65),
        Bookmark("Chapter 5: Conclusion", 70),
        Bookmark("Appendix", 75),
        Bookmark("Index", 80),
        Bookmark("End of Book", 85) # A virtual bookmark to define the end of the last actual content chapter
    ]
    placeholder_total_pages = 85 # Total pages in the "book"

    # --- Configuration for Reading Plan ---
    desired_total_days = 10
    plan_start_date = datetime(2023, 11, 1)
    take_weekends_off = True # Set to False if you want to read on weekends

    analyzer = ChapterWeightAnalyzer(bookmarks=placeholder_bookmarks, total_pages=placeholder_total_pages)
    chapter_weights_data = analyzer.get_level1_weights() # Use level 1 for main chapters for the plan

    # Sort chapter_weights_data by start_page to ensure correct order for grouping
    chapter_weights_data.sort(key=lambda cw: cw.start_page)

    plan_generator = ReadingPlanGenerator(
        chapter_weights=chapter_weights_data,
        total_target_days=desired_total_days,
        start_date=plan_start_date,
        weekends_off=take_weekends_off
    )
    reading_plan = plan_generator.generate_plan()

    # --- Print the generated plan ---
    print("Generated Reading Plan:\n")
    print(f"{'Reading Block':<40} {'Pages':<8} {'Days':<5} {'Start Date':<12} {'End Date':<12}")
    print(f"{'-'*40:<40} {'-'*8:<8} {'-'*5:<5} {'-'*12:<12} {'-'*12:<12}")

    for entry in reading_plan:
        print(f"{entry.chapter_title:<40} {entry.page_count:<8} {entry.assigned_days:<5} {entry.start_date:<12} {entry.end_date:<12}")
        if entry.merged_chapter_titles:
             print(f"{'':<4} (Includes: {', '.join(entry.merged_chapter_titles)})")

    print(f"\nTotal Target Days: {desired_total_days}")
    print(f"Actual Assigned Days: {sum(entry.assigned_days for entry in reading_plan)}")
    print(f"Average Daily Pages (Active Days): {plan_generator.get_daily_pages():.1f}")

    # --- Export options ---
    output_base_name = "smart_reading_plan"

    if OPENPYXL_AVAILABLE:
        excel_filepath = f"{output_base_name}.xlsx"
        if plan_generator.export_to_excel(excel_filepath):
            print(f"\nReading plan exported to {excel_filepath}")
        else:
            print(f"\nFailed to export reading plan to Excel.")

    markdown_filepath = f"{output_base_name}.md"
    if plan_generator.export_to_markdown(markdown_filepath):
        print(f"Reading plan exported to {markdown_filepath}")
    else:
        print(f"Failed to export reading plan to Markdown.")

    json_filepath = f"{output_base_name}.json"
    if plan_generator.export_to_json(json_filepath):
        print(f"Reading plan exported to {json_filepath}")
    else:
        print(f"Failed to export reading plan to JSON.")

    text_filepath = f"{output_base_name}.txt"
    if plan_generator.export_to_text(text_filepath):
        print(f"Reading plan exported to {text_filepath}")
    else:
        print(f"Failed to export reading plan to Text.")