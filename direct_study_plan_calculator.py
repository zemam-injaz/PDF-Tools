#!/usr/bin/env python3
"""
Direct Study Plan Calculator

A simpler, predictable, and mathematically sound algorithm that maps chapter
pages directly to days with a perfect linear relationship.

Key ideas:
- days_for_chapter = (chapter_pages / total_pages) * total_target_days
- Prefer Largest Remainder Method (Hamilton) for rounding so the sum of days
  equals exactly total_target_days
- When there are enough days (total_target_days >= number_of_chapters), guarantee
  at least 1 day per chapter without merging
- Optional weekend skipping when assigning calendar dates

This file does NOT modify or depend on weight percentages.
"""

from __future__ import annotations

from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timedelta
from dataclasses import asdict

# Reuse the same data classes to keep the interface identical
try:
    # Same-folder import when used inside the app
    from chapter_weight_analyzer import ChapterWeight, ReadingPlanEntry  # type: ignore
except Exception:
    # Fallback for relative import contexts
    from .chapter_weight_analyzer import ChapterWeight, ReadingPlanEntry  # type: ignore

# Optional excel support (used by export methods)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl.comments
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


class DirectStudyPlanCalculator:
    """
    Create a study plan where allocated days are directly proportional to
    chapter page counts.

    Interface compatibility:
    - Accepts List[ChapterWeight]
    - Returns List[ReadingPlanEntry]
    - Supports total_target_days, start_date, weekends_off, weekend_days
    """

    def __init__(
        self,
        chapter_weights: List[ChapterWeight],
        total_target_days: int,
        start_date: Optional[datetime] = None,
        weekends_off: bool = False,
        weekend_days: Optional[List[int]] = None,
    ) -> None:
        # Keep only chapters with positive page counts
        self.chapters: List[ChapterWeight] = [c for c in chapter_weights if c.page_count > 0]
        self.total_target_days = int(max(0, total_target_days))
        self.start_date = start_date or datetime.now()
        self.weekends_off = weekends_off
        self.weekend_days = weekend_days if weekend_days is not None else [5, 6]  # Sat, Sun
        self.reading_plan: List[ReadingPlanEntry] = []

    # ------------------------- Core allocation logic ------------------------- #

    def _largest_remainder_round(self, raw_days: List[float], total_days: int) -> List[int]:
        """Round a list of raw day values so their integer sum equals total_days.

        Uses Hamilton/Largest Remainder method. Ties broken by larger chapters (implicitly
        by their raw remainder order) and then by original index to keep stable order.
        """
        floors = [int(x) for x in raw_days]
        remainders = [x - int(x) for x in raw_days]
        current_sum = sum(floors)
        needed = total_days - current_sum

        if needed <= 0:
            # Nothing to add; if negative due to numerics, trim from smallest remainders
            if needed < 0:
                # remove |needed| units from the smallest remainders, but never below 0
                pairs = sorted([(r, i) for i, r in enumerate(remainders)])
                for _, i in pairs[:abs(needed)]:
                    if floors[i] > 0:
                        floors[i] -= 1
            return floors

        # Distribute remaining days to chapters with the largest remainders
        # Break ties preferring larger raw value (more pages), then lower index (stable)
        candidates = sorted(
            [(remainders[i], raw_days[i], i) for i in range(len(raw_days))],
            key=lambda t: (t[0], t[1]),
            reverse=True,
        )
        for _, __, idx in candidates[:needed]:
            floors[idx] += 1
        return floors

    def _allocate_days(self) -> List[int]:
        """Allocate integer days for each chapter based purely on pages.

        Two modes:
        - If we have at least as many days as chapters, guarantee >= 1 day/chapter
          by giving a baseline of 1 day and distributing the remaining days by LRM.
        - Otherwise (days < chapters), use standard LRM; some small chapters may get 0 days.
        """
        n = len(self.chapters)
        if n == 0 or self.total_target_days <= 0:
            return []

        total_pages = sum(ch.page_count for ch in self.chapters)
        if total_pages <= 0:
            return []

        # Raw proportional days (perfect linear relation to pages)
        raw = [
            (ch.page_count / total_pages) * self.total_target_days for ch in self.chapters
        ]

        # Case A: enough days to ensure at least 1 day per chapter
        if self.total_target_days >= n:
            # Give each chapter 1 baseline day, then allocate the remainder proportionally
            remaining_days = self.total_target_days - n
            # Effective extra over the baseline 1 day
            raw_extra = [max(0.0, x - 1.0) for x in raw]
            extra = self._largest_remainder_round(raw_extra, remaining_days)
            return [1 + e for e in extra]

        # Case B: fewer days than chapters; allocate proportionally, some may be 0
        return self._largest_remainder_round(raw, self.total_target_days)

    # --------------------------- Calendar assignment ------------------------- #

    def _skip_to_next_reading_day(self, d: datetime) -> datetime:
        if not self.weekends_off:
            return d
        while d.weekday() in self.weekend_days:
            d += timedelta(days=1)
        return d

    def _add_reading_days(self, start: datetime, days: int) -> datetime:
        """Return the end date after counting 'days' reading days from start."""
        if days <= 0:
            return start
        count = 0
        d = start
        while count < days:
            d = self._skip_to_next_reading_day(d)
            count += 1
            if count < days:
                d += timedelta(days=1)
        return d

    # ------------------------------- Public API ------------------------------ #

    def generate_plan(self) -> List[ReadingPlanEntry]:
        """Generate the reading plan entries in chronological order.

        - Ensures sum(assigned_days) == total_target_days
        - When possible (days >= chapters), every chapter gets at least one day
        - Dates respect weekend skipping if enabled
        """
        self.reading_plan = []
        if not self.chapters or self.total_target_days <= 0:
            return []

        days_per_chapter = self._allocate_days()
        if not days_per_chapter:
            return []

        current = self._skip_to_next_reading_day(self.start_date)

        for ch, days in zip(self.chapters, days_per_chapter):
            if days <= 0:
                # Not enough total days to cover every chapter; mark with no dates
                entry = ReadingPlanEntry(
                    chapter_title=ch.title,
                    start_page=ch.start_page,
                    end_page=ch.end_page,
                    page_count=ch.page_count,
                    weight_percentage=getattr(ch, 'weight_percentage', 0.0),
                    assigned_days=0,
                    start_date=None,
                    end_date=None,
                    merged_chapter_titles=None,
                )
                self.reading_plan.append(entry)
                continue

            start = self._skip_to_next_reading_day(current)
            end = self._add_reading_days(start, days)

            entry = ReadingPlanEntry(
                chapter_title=ch.title,
                start_page=ch.start_page,
                end_page=ch.end_page,
                page_count=ch.page_count,
                weight_percentage=getattr(ch, 'weight_percentage', 0.0),
                assigned_days=days,
                start_date=start.strftime('%Y-%m-%d'),
                end_date=end.strftime('%Y-%m-%d'),
                merged_chapter_titles=None,
            )
            self.reading_plan.append(entry)

            # next chapter begins the day after end
            current = end + timedelta(days=1)

        return self.reading_plan

    def get_average_pages_per_day(self) -> float:
        total_pages = sum(ch.page_count for ch in self.chapters)
        return (total_pages / self.total_target_days) if self.total_target_days > 0 else 0.0

    def get_statistics(self) -> Dict[str, Any]:
        if not self.reading_plan:
            return {}
        return {
            'total_chapters': len(self.reading_plan),
            'total_pages': sum(e.page_count for e in self.reading_plan),
            'total_days': sum(e.assigned_days for e in self.reading_plan),
            'avg_pages_per_day': self.get_average_pages_per_day(),
            'min_chapter_days': min(e.assigned_days for e in self.reading_plan),
            'max_chapter_days': max(e.assigned_days for e in self.reading_plan),
        }


    def get_daily_pages(self) -> float:
        """Average pages per actual assigned reading day."""
        if not self.reading_plan:
            return 0.0
        total_pages = sum(entry.page_count for entry in self.reading_plan)
        active_days = sum(entry.assigned_days for entry in self.reading_plan)
        return (total_pages / active_days) if active_days > 0 else 0.0

    # ------------------------------- Exports --------------------------------- #

    def export_to_dict(self) -> List[Dict[str, Any]]:
        return [asdict(entry) for entry in self.reading_plan]

    def export_to_json(self, filepath: str) -> bool:
        try:
            import json
            data = {
                'total_target_days': self.total_target_days,
                'actual_assigned_days': sum(entry.assigned_days for entry in self.reading_plan),
                'start_date': self.start_date.strftime('%Y-%m-%d'),
                'average_daily_pages': self.get_daily_pages(),
                'plan_created': datetime.now().isoformat(),
                'reading_plan': self.export_to_dict(),
            }
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"Error exporting reading plan to JSON: {e}")
            return False

    def export_to_csv(self, filepath: str) -> bool:
        try:
            import csv
            if not self.reading_plan:
                return False
            fieldnames = ['chapter_title', 'start_page', 'end_page', 'page_count',
                          'weight_percentage', 'assigned_days', 'start_date', 'end_date', 'merged_chapter_titles']
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for entry in self.reading_plan:
                    row = asdict(entry)
                    if row['merged_chapter_titles'] is not None:
                        row['merged_chapter_titles'] = '; '.join(row['merged_chapter_titles'])
                    writer.writerow(row)
            return True
        except Exception as e:
            print(f"Error exporting reading plan to CSV: {e}")
            return False

    def export_to_text(self, filepath: str, localization=None) -> bool:
        try:
            # Helper function to get localized text
            def get_text(key: str, fallback: str = "") -> str:
                if localization and hasattr(localization, 'get_text'):
                    return localization.get_text(key)
                return fallback if fallback else key

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write(get_text("report_reading_plan", "READING PLAN") + "\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"{get_text('report_total_target_duration', 'Total Target Duration')}: {self.total_target_days} {get_text('report_days', 'days')}\n")
                f.write(f"{get_text('report_actual_assigned_duration', 'Actual Assigned Duration')}: {sum(e.assigned_days for e in self.reading_plan)} {get_text('report_days', 'days')}\n")
                f.write(f"{get_text('report_start_date', 'Start Date')}: {self.start_date.strftime('%Y-%m-%d')}\n")
                f.write(f"{get_text('report_average_daily_pages', 'Average Daily Pages (Active Days)')}: {self.get_daily_pages():.1f}\n\n")
                f.write("-" * 80 + "\n\n")
                current_reading_day_counter = 1
                for i, entry in enumerate(self.reading_plan, 1):
                    days_label = get_text('report_days', 'Days')
                    f.write(f"{get_text('report_block', 'Block')} {i} ({days_label} {current_reading_day_counter}-{current_reading_day_counter + entry.assigned_days - 1}) - {entry.start_date} {get_text('report_to', 'to')} {entry.end_date}\n")
                    f.write(f"{get_text('report_chapter', 'Chapter')}: {entry.chapter_title}\n")
                    if entry.merged_chapter_titles:
                        f.write(f"  ({get_text('report_includes', 'Includes')}: {', '.join(entry.merged_chapter_titles)})\n")
                    f.write(f"{get_text('report_pages', 'Pages')}: {entry.start_page}-{entry.end_page} ({entry.page_count} {get_text('report_pages', 'pages')})\n")
                    f.write(f"{get_text('report_duration', 'Duration')}: {entry.assigned_days} {get_text('report_days_unit', 'day(s)')}\n")
                    f.write(f"{get_text('report_weight', 'Weight')}: {entry.weight_percentage:.2f}%\n\n")
                    current_reading_day_counter += entry.assigned_days
                f.write("=" * 80 + "\n")
                f.write(f"{get_text('report_plan_created', 'Plan created')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            return True
        except Exception as e:
            print(f"Error exporting reading plan to text: {e}")
            return False

    def export_to_excel(self, filepath: str, localization=None) -> bool:
        if not OPENPYXL_AVAILABLE:
            print("Error: openpyxl is required for Excel export")
            return False
        try:
            # Helper function to get localized text
            def get_text(key: str, fallback: str = "") -> str:
                if localization and hasattr(localization, 'get_text'):
                    return localization.get_text(key)
                return fallback if fallback else key

            wb = Workbook()
            ws = wb.active
            ws.title = get_text("report_reading_plan", "Reading Plan")

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(horizontal="left", vertical="center")
            number_alignment = Alignment(horizontal="right", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Summary
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = get_text("report_reading_plan", "Reading Plan")
            title_cell.font = Font(bold=True, size=14, color="1F4E78")
            title_cell.alignment = center_alignment

            ws['A2'] = get_text("report_total_target_duration", "Total Target Duration") + ":"
            ws['C2'] = f"{self.total_target_days} {get_text('report_days', 'days')}"
            ws['A3'] = get_text("report_actual_assigned_duration", "Actual Assigned Duration") + ":"
            ws['C3'] = f"{sum(e.assigned_days for e in self.reading_plan)} {get_text('report_days', 'days')}"
            ws['A4'] = get_text("report_start_date", "Start Date") + ":"
            ws['C4'] = self.start_date.strftime('%Y-%m-%d')
            ws['A5'] = get_text("report_average_daily_pages", "Average Daily Pages (Active Days)") + ":"
            ws['C5'] = f"{self.get_daily_pages():.1f}"

            headers = [
                get_text("chapter_title", "Chapter Title"),
                get_text("start_page", "Start Page"),
                get_text("end_page", "End Page"),
                get_text("page_count", "Page Count"),
                get_text("weight_percent", "Weight %"),
                get_text("assigned_days", "Days"),
                get_text("start_date", "Start Date"),
                get_text("end_date", "End Date")
            ]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row_num, entry in enumerate(self.reading_plan, 8):
                title_cell = ws.cell(row=row_num, column=1, value=entry.chapter_title)
                title_cell.alignment = cell_alignment
                if entry.merged_chapter_titles:
                    title_cell.comment = openpyxl.comments.Comment(
                        f"{get_text('report_includes', 'Includes')}: {', '.join(entry.merged_chapter_titles)}\n"
                        f"{get_text('total_pages', 'Total')}: {entry.page_count} {get_text('report_pages', 'pages')}, "
                        f"{get_text('report_duration', 'Duration')}: {entry.assigned_days} {get_text('report_days', 'days')}",
                        author="DirectStudyPlanCalculator",
                    )
                ws.cell(row=row_num, column=2, value=entry.start_page).alignment = number_alignment
                ws.cell(row=row_num, column=3, value=entry.end_page).alignment = number_alignment
                ws.cell(row=row_num, column=4, value=entry.page_count).alignment = number_alignment
                ws.cell(row=row_num, column=5, value=f"{entry.weight_percentage:.2f}%").alignment = number_alignment
                ws.cell(row=row_num, column=6, value=entry.assigned_days).alignment = number_alignment
                ws.cell(row=row_num, column=7, value=entry.start_date).alignment = center_alignment
                ws.cell(row=row_num, column=8, value=entry.end_date).alignment = center_alignment
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).border = border

            column_widths = [40, 12, 12, 12, 10, 8, 14, 14]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            ws.freeze_panes = 'A8'
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Error exporting reading plan to Excel: {e}")
            return False

    def export_to_markdown(self, filepath: str, localization=None) -> bool:
        try:
            # Helper function to get localized text
            def get_text(key: str, fallback: str = "") -> str:
                if localization and hasattr(localization, 'get_text'):
                    return localization.get_text(key)
                return fallback if fallback else key

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("---\n")
                f.write("tags: [reading, plan, schedule]\n")
                f.write(f"created: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write(f"start_date: {self.start_date.strftime('%Y-%m-%d')}\n")
                f.write(f"duration_target: {self.total_target_days}\n")
                f.write(f"duration_actual: {sum(e.assigned_days for e in self.reading_plan)}\n")
                f.write("type: reading-plan\n")
                f.write("---\n\n")

                f.write(f"# {get_text('report_reading_plan', 'Reading Plan')}\n\n")
                f.write(f"## {get_text('plan_summary', 'Plan Summary')}\n\n")
                f.write(f"- **{get_text('report_total_target_duration', 'Total Target Duration')}**: {self.total_target_days} {get_text('report_days', 'days')}\n")
                f.write(f"- **{get_text('report_actual_assigned_duration', 'Actual Assigned Duration')}**: {sum(e.assigned_days for e in self.reading_plan)} {get_text('report_days', 'days')}\n")
                f.write(f"- **{get_text('report_start_date', 'Start Date')}**: {self.start_date.strftime('%Y-%m-%d')}\n")
                if self.reading_plan:
                    f.write(f"- **{get_text('end_date', 'End Date')}**: {self.reading_plan[-1].end_date}\n")
                f.write(f"- **{get_text('report_average_daily_pages', 'Average Daily Pages (Active Days)')}**: {self.get_daily_pages():.1f}\n")
                f.write(f"- **{get_text('total_reading_blocks', 'Total Reading Blocks')}**: {len(self.reading_plan)}\n\n")

                f.write(f"## {get_text('reading_schedule', 'Reading Schedule')}\n\n")
                current_day_overall = 1
                for entry in self.reading_plan:
                    f.write(f"### {entry.chapter_title}\n\n")
                    if entry.merged_chapter_titles:
                        f.write(f"  > **{get_text('report_includes', 'Includes')}**: {', '.join(entry.merged_chapter_titles)}\n\n")
                    f.write(f"- **{get_text('report_pages', 'Pages')}**: {entry.start_page}-{entry.end_page} ({entry.page_count} {get_text('report_pages', 'pages')})\n")
                    f.write(f"- **{get_text('report_duration', 'Duration')}**: {entry.assigned_days} {get_text('report_days_unit', 'day(s)')}\n")
                    f.write(f"- **{get_text('dates', 'Dates')}**: {entry.start_date} → {entry.end_date}\n")
                    f.write(f"- **{get_text('report_weight', 'Weight')}**: {entry.weight_percentage:.2f}%\n")

                    days_label = get_text('report_days', 'Days')
                    f.write(f"\n**{get_text('daily_progress', 'Daily Progress')} ({get_text('reading_days', 'Reading Days')} {current_day_overall} {get_text('report_to', 'to')} {current_day_overall + entry.assigned_days - 1})**:\n")
                    for day_offset in range(entry.assigned_days):
                        temp_date = datetime.strptime(entry.start_date, '%Y-%m-%d')
                        # advance by day_offset respecting weekends
                        for _ in range(day_offset):
                            temp_date += timedelta(days=1)
                            while self.weekends_off and temp_date.weekday() in self.weekend_days:
                                temp_date += timedelta(days=1)
                        f.write(f"- [ ] {get_text('report_day', 'Day')} {current_day_overall + day_offset} ({temp_date.strftime('%Y-%m-%d')})\n")
                    current_day_overall += entry.assigned_days
                    f.write("\n")

                f.write(f"## {get_text('quick_reference', 'Quick Reference')}\n")
                lines = []
                lines.append(f"| {get_text('reading_block', 'Reading Block')} | {get_text('report_pages', 'Pages')} | {get_text('report_days', 'Days')} | {get_text('start_date', 'Start Date')} | {get_text('end_date', 'End Date')} |")
                lines.append("|:--------------|-----:|----:|:----------:|:--------:|")
                for entry in self.reading_plan:
                    title = entry.chapter_title.replace('|', '\\|').replace('\n', ' ')
                    lines.append(f"| {title} | {entry.page_count} | {entry.assigned_days} | {entry.start_date} | {entry.end_date} |")
                f.write("\n".join(lines))
                f.write("\n\n---\n")
                f.write(f"*{get_text('generated', 'Generated')}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")
            return True
        except Exception as e:
            print(f"Error exporting reading plan to Markdown: {e}")
            return False


    def export_to_obsidian_markdown(self, schedule, filename: str = "reading_plan.md") -> bool:
        """Export reading plan to Obsidian-friendly Markdown.

        Structure:
        - YAML front matter (tags, created, type)
        - Title, Summary (based on statistics)
        - Reading schedule with per-day checkboxes
        - Quick reference table (with compact month names)
        """
        try:
            if schedule is None:
                schedule = self.reading_plan
            if not schedule:
                return False

            stats = self.get_statistics(schedule)

            with open(filename, 'w', encoding='utf-8') as f:
                # Front matter
                f.write("---\n")
                f.write("tags: [reading, plan, schedule]\n")
                f.write(f"created: {datetime.now().strftime('%Y-%m-%d')}\n")
                f.write("type: reading-plan\n")
                f.write("---\n\n")

                # Title
                f.write("# Reading Plan\n\n")

                # Summary
                f.write("## Plan Summary\n\n")
                f.write(f"- **Total Days**: {stats.get('total_days', 0)}\n")
                f.write(f"- **Total Pages**: {stats.get('total_pages', 0)}\n")
                f.write(f"- **Daily Average**: {stats.get('avg_pages_per_day', 0)} pages\n")
                f.write(f"- **Total Blocks**: {stats.get('total_sessions', len(schedule))}\n\n")
                f.write("---\n\n")

                # Schedule
                f.write("## Reading Schedule\n\n")

                for i, session in enumerate(schedule, 1):
                    f.write(f"### Block {i}: {session.chapter_title}\n\n")
                    f.write(f"**Pages**: {session.start_page}-{session.end_page} ({session.page_count} pages)  \n")
                    f.write(f"**Duration**: {session.assigned_days} day{'s' if session.assigned_days > 1 else ''}  \n")
                    f.write(f"**Dates**: {session.start_date} → {session.end_date}  \n\n")

                    # Checkboxes for each day
                    f.write("**Daily Progress**:\n")
                    current = datetime.strptime(session.start_date, "%Y-%m-%d")
                    end = datetime.strptime(session.end_date, "%Y-%m-%d")
                    day_num = sum(s.assigned_days for s in schedule[:i-1]) + 1

                    while current <= end:
                        f.write(f"- [ ] Day {day_num} ({current.strftime('%Y-%m-%d')})\n")
                        day_num += 1
                        current += timedelta(days=1)

                    f.write("\n---\n\n")

                # Quick table
                f.write("## Quick Reference\n\n")
                f.write("| Block | Pages | Days | Start | End |\n")
                f.write("|:------|------:|-----:|:-----:|:---:|\n")
                for i, s in enumerate(schedule, 1):
                    title = s.chapter_title[:30] + "..." if len(s.chapter_title) > 30 else s.chapter_title
                    start = datetime.strptime(s.start_date, "%Y-%m-%d").strftime("%b %d")
                    end = datetime.strptime(s.end_date, "%Y-%m-%d").strftime("%b %d")
                    f.write(f"| {title} | {s.page_count} | {s.assigned_days} | {start} | {end} |\n")

            return True
        except Exception as e:
            print(f"Error exporting to Obsidian markdown: {e}")
            return False
